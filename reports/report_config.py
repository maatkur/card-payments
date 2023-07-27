from os import getlogin, system

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from helpers.date_helpers import DateHelpers


class ReportConfig:
    def __init__(self, header, report_name):
        self.header = header
        self.report_name = report_name
        self.workbook = openpyxl.Workbook()  # Init workbook
        self.worksheet = self.workbook.active  # Init worksheet
        if type(self) is ReportConfig:
            raise NotImplementedError("ReportConfig class cannot be started directly, you must use it as inheritance")

    def create_new_worksheet(self, sheet_name: str) -> None:
        self.workbook.create_sheet(sheet_name)

    def switch_worksheet(self, sheet_name: str) -> None:
        self.worksheet = self.workbook[sheet_name]

    def set_header_style(self) -> None:
        # Define a cor das células do cabeçalho como azul claro
        header_fill = PatternFill(start_color="00CCFF", end_color="00CCFF", fill_type="solid")

        # Define o estilo do cabeçalho em negrito e cor da letra branca
        bold_font = Font(bold=True, color="FFFFFF")

        for sheet in self.workbook.sheetnames:
            worksheet = self.workbook[sheet]
            for col_num, cell in enumerate(worksheet[1], 1):
                col_letter = get_column_letter(col_num)
                header_cell = worksheet[f"{col_letter}1"]
                header_cell.fill = header_fill
                header_cell.font = bold_font

                # Definir largura da coluna com base no tamanho do conteúdo
                worksheet.column_dimensions[col_letter].auto_size = True

    def save_workbook(self) -> None:
        # Salva o arquivo temporário
        user = getlogin()
        date = DateHelpers.to_date_string()
        temp_filename = rf'C:\Users\{user}\Desktop\{self.report_name}{date}.xlsx'
        self.workbook.save(temp_filename)
        system(f'start excel.exe "{temp_filename}"')

    def reset(self) -> None:
        self.workbook = openpyxl.Workbook()  # Init workbook
        self.worksheet = self.workbook.active
